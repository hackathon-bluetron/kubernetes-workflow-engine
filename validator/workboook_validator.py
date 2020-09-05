#=====================================================================================#
#title           :workbook_validator.py                                               #
#description     :This script validates the uploadded workbook data                   #
#author          :Pravat B                                                            #
#date            :20200905                                                            #
#version         :0.8                                                                 #
#usage           :py workbook_validator.py -i <<name-of-file-to-process>>             #
#notes           :N/A                                                                 #
#python_version  :3.6                                                                 #
#=====================================================================================#



import openpyxl
import argparse
import os
import numpy as np

parser = argparse.ArgumentParser(prog='workbook_validator.py', description='%(prog)s - Validate the workbook after removing empty rows')
parser.add_argument("-i", "--input", dest='fname')
args = parser.parse_args()

FNAME = args.fname

if not FNAME:
    print ("python workbook_validator.py -i <name/of/workbook>")
    exit(1)

wb = openpyxl.load_workbook(FNAME)
res = len(wb.sheetnames)
workdir = (str('/mnt' + '/data' + '/input'))
os.chdir(workdir)

for sheet in wb.worksheets:
    max_row_in_sheet = sheet.max_row
    max_col_in_sheet = sheet.max_column
    array_3 = np.array([])
    array_4 = np.array([])
    r = 1
    c = 1
    for r in range(1, max_row_in_sheet+1):
        array_1 = np.array([])
        array_2 = np.array([])
        for c in range (1, max_col_in_sheet+1):
            if sheet.cell(row = r, column = c).value == None:
                array_1 = np.append(array_2, c)
                array_2 = array_1
        if len(array_1) == max_col_in_sheet:
            array_3 = np.append(array_4, r)
            array_4 = array_3
            array_3 = array_3.astype(int)
    if len(array_3) != 0:
        index_of_last_array_element = len(array_3) - 1
        while index_of_last_array_element != -1:
            sheet.delete_rows(array_3[index_of_last_array_element], 1)
            index_of_last_array_element = index_of_last_array_element - 1
wb.save(FNAME)
